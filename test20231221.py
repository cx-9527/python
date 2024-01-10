# -*- codeing = utf-8 -*-
from bs4 import BeautifulSoup  # 网页解析，获取数据
import re  # 正则表达式，进行文字匹配`
import urllib.request, urllib.error  # 制定URL，获取网页数据
import random
import urllib.parse
import ua_info
import time
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

# 创建excel文件
book_1 = openpyxl.Workbook()
book_1.save('D:/python/Projects/Pc/test_4-2022/test01.xlsx')
book_1.close()
time.sleep(1)

# 电影链接
findLink = re.compile(r'<a href="(.*?)">')
# 图片
findsrc = re.compile(r'src="(.*?)"', re.S)
# 片名
findname = re.compile(r'<span class="title">(.*)</span>')
# 评分
findscore = re.compile(r'<span class="rating_num" property="v:average">(.*)</span>')
# 评价数
findpj = re.compile(r'<span>(.*)</span>')
# 概况
findgk = re.compile(r'<span class="inq">(.*)</span>')
# 相关信息
findmore = re.compile(r'<p class="">(.*?)</p>', re.S)
# 定义url与headers
url = 'https://movie.douban.com/top250?start={}'
# 创建代理池，随机选取ua
chooseus = random.choice(ua_info.ua_list)
# 重构请求头，伪装ua
headers = {
    'User-Agent': chooseus
}
'''
headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.12; rv:65.0) Gecko/20100101 Firefox/65.0'
}
'''
# 创建列表
datalist = []
# 进行多个页面请求
for i in range(0, 6):
    url1 = url.format(i*25)
    print(url1)
    rq = urllib.request.Request(url=url1, headers=headers)
    # 发送请求
    response = urllib.request.urlopen(rq)
    time.sleep(1)
    # 提取响应内容
    html = response.read().decode("utf-8")
    time.sleep(1)
    soup = BeautifulSoup(html, "html.parser")
    time.sleep(1)
    test = soup.find_all('div', class_="item")
    for k in test:
        data = []
        item = str(k)
        # 获取链接
        test1 = re.findall(findLink, item)[0]
        data.append(test1)
        time.sleep(1)
        # 获取图片
        test2 = re.findall(findsrc, item)[0]
        data.append(test2)
        time.sleep(1)
        test3 = re.findall(findname, item)
        if len(test3) == 2:
            # 获取电影片名（中文）
            data.append(test3[0])
            # 获取电影片名（外国名）
            # 消除转义符
            name = test3[1].replace("/", "")
            # 移除字符串头尾指定的字符
            name_wai = name.strip('\xa0\xa0')
            data.append(name_wai)
        else:
            data.append(test3[0])
            data.append('咩有！')
        time.sleep(1)
        # 获取评分
        test4 = re.findall(findscore, item)[0]
        data.append(test4)
        time.sleep(1)
        # 获取评论数
        test5 = re.findall(findpj, item)[0]
        data.append(test5)
        time.sleep(1)
        # 获取概况
        test6 = re.findall(findgk, item)
        # 判断概况是否为空
        if len(test6) == 1:
            data.append(test6[0])
        else:
            data.append('咩有！')
        time.sleep(1)
        # 获取相关信息
        test7 = re.findall(findmore, item)
        more = test7[0].replace("\n", "")
        more_z1 = more.strip('\xa0/\xa0')
        more_z2 = more_z1.strip()
        more_z3 = more_z2.strip('...<br/>')
        data.append(more_z3)
        datalist.append(data)
print(datalist)



# 加载excel文件
book_2 = openpyxl.load_workbook('D:/python/Projects/Pc/test_4-2022/test01.xlsx')
# 创建工作表
sheet = book_2.create_sheet(index=0, title="电影TOP")
book_2.save('D:/python/Projects/Pc/test_4-2022/test01.xlsx')
book_2.close()
time.sleep(1)
book_2 = openpyxl.load_workbook('D:/python/Projects/Pc/test_4-2022/test01.xlsx')
# 获取当前工作表
w_s = book_2.active
# 定义表头
col = ("电影详情链接", "图片链接", "影片中文名", "影片外国名", "评分", "评价数", "概况", "相关信息")
for i in range(1, 9):
    w_s.cell(1, i, col[i-1])

for c in range(1, len(datalist)+1):
    data1 = datalist[c-1]
    for j in range(1, 9):
        w_s.cell(c+1, j, data1[j-1])

# 设置列的宽度
w_s.column_dimensions["A"].width = 25
w_s.column_dimensions["B"].width = 30
w_s.column_dimensions["C"].width = 20
w_s.column_dimensions["D"].width = 15
w_s.column_dimensions["E"].width = 15
w_s.column_dimensions["F"].width = 15
w_s.column_dimensions["G"].width = 25
w_s.column_dimensions["H"].width = 30
# 表内容居中
for row in w_s.iter_rows(min_row=2, min_col=1):
    for cell in row:
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
book_2.save('D:/python/Projects/Pc/test_4-2022/test01.xlsx')

'''
# 创建请求对象，包装ua信息
rq = urllib.request.Request(url=url, headers=headers)
# 发送请求
response = urllib.request.urlopen(rq)
# 提取响应内容
html = response.read().decode()
print(html)
'''
